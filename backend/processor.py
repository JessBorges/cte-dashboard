"""Data processing pipeline for CTE enrollment data.

Extracted from build_tracker.py — matrix parsing, enrollment parsing,
seat allocation parsing, and enrollment resolution.
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

from campus_norm import enrollment_key_from_sis

RESOURCES = Path(__file__).resolve().parent.parent.parent / "Resources"
MATRIX_FILE = RESOURCES / "Dallas ISD CTE Programs Matrix - CTE Matrix w_ NAF.csv"
SEATS_FILE = RESOURCES / "CTE_Career Institutes PoS - Dallas ISD - Information for 26-27.xlsx"
POS_NOTES_FILE = RESOURCES / "25-26 Programs Of Study Notes (Internal).xlsx"

# Notes workbook sheet tab → enrollment matrix campus key
NOTES_SHEET_TO_CAMPUS: dict[str, str] = {
    "Bryan Adams": "Bryan Adams",
    "Adamson": "Adamson",
    "New Tech": "Smith New Tech",
    "Kimball": "Kimball",
    "Molina": "Molina",
    "Hillcrest": "Hillcrest",
    "Lincoln": "Lincoln",
    "TJ": "Jefferson",
    "Pinkston": "Pinkston",
    "Roosevelt": "Roosevelt",
    "Sunset": "Sunset",
    "Samuell": "Samuell",
    "Townview - TAG": "School for Talented & Gifted",
    "SOC": "South Oak Cliff",
    "Spruce": "Spruce",
    "Seagoville": "Seagoville",
    "White": "WT White",
    "Wilson": "Woodrow Wilson",
    "Carter": "Carter",
    "North Dallas": "North Dallas",
    "Skyline": "Skyline",
    "Townview SEM": "School of Science & Engineering",
    "Conrad": "Conrad",
    "Madison": "Madison",
    "Townview - Bus": "School of Business/Management",
    "Irma Rangel": "Rangel YWLS",
    "Townview - Health": "School of Health Professions",
    "Townview - Education": "Sorrells Education/Social Services",
    "Townview - Law": "Sanders Public Service, Law",
    "IDEA": "IDEA",
    "Wilmer-Hutchins": "Wilmer-Hutchins",
    "Gilliam": "Gilliam",
    "Obama": "Obama",
    "City Lab": "City Lab",
}

# ── First course(s) per PoS ─────────────────────────────────────────

FIRST_COURSES: dict[str, list[int]] = {
    "Animal Science": [7238],
    "Applied Agricultural Engineering - Welding": [7238],
    "Environmental & Natural Resources": [7238],
    "Floral Design": [7238],
    "Horticulture Science": [7238],
    "Architectural Design": [7023],
    "Carpentry": [7021],
    "Construction Management & Inspection": [7021],
    "Electrical": [7021],
    "HVAC & Sheet Metal": [7021],
    "Interior Design": [7023],
    "Plumbing & Pipe Fitting": [7021],
    "Animation": [7355],
    "Digital Communications": [7355],
    "Fashion Design": [7355],
    "Graphic Design": [7355],
    "Photography": [7355],
    "Video Game Design": [7071],
    "Accounting & Financial Services": [7154, 7151],
    "Business Management": [7154, 7151],
    "Entrepreneurship": [7154, 7151],
    "Marketing & Sales": [7315, 7151],
    "Teaching & Training": [7412],
    "Exercise Science & Wellness": [7426],
    "Health Informatics - Medical Coding & Billing": [7426],
    "Healthcare Diagnostic - ECG/EKG & Phlebotomy": [7426],
    "Biomedical Science": [7716],
    "Healthcare Therapeutic - Patient Care Tech": [7426],
    "Healthcare Therapeutic - Pharmacy Technician": [7426],
    "Healthcare Therapeutic - Registered Dental Asst.": [7426],
    "Healthcare Therapeutic - Emergency Med. Tech.": [7426],
    "Culinary Arts": [7803, 7270],
    "Lodging & Resort Management": [7270],
    "Travel, Tourism & Attractions": [7270],
    "Cosmetology & Personal Care": [7303],
    "Family & Community Services": [7291],
    "Health & Wellness": [7291],
    "Cybersecurity": [2710, 2690],
    "Information Technology Support & Services": [7309],
    "Networking Systems": [7309],
    "Programming & Software Development": [2690],
    "Web Development": [7315, 7309],
    "Emergency Services - Firefighter": [7326],
    "Government & Public Administration": [7326],
    "Law Enforcement": [7326],
    "Legal Studies": [7326],
    "Electronics Technology": [7175],
    "Welding": [7824],
    "Engineering Foundations": [7175],
    "Automotive Collision": [7857],
    "Automotive Maintenance": [7850],
    "Aviation Maintenance": [7854],
    "Aviation Flight": [3112, 7854],
}

CI_PROGRAMS = {
    "Architectural Design", "Carpentry", "Construction Management & Inspection",
    "Electrical", "HVAC & Sheet Metal", "Interior Design", "Plumbing & Pipe Fitting",
    "Healthcare Diagnostic - ECG/EKG & Phlebotomy",
    "Healthcare Therapeutic - Patient Care Tech",
    "Healthcare Therapeutic - Pharmacy Technician",
    "Healthcare Therapeutic - Registered Dental Asst.",
    "Healthcare Therapeutic - Emergency Med. Tech.",
    "Cybersecurity", "Programming & Software Development",
    "Electronics Technology", "Welding",
    "Automotive Collision", "Automotive Maintenance",
    "Aviation Maintenance", "Aviation Flight",
}

# ── Campus name map (shared with IBC via campus_norm) ─────────────────


def _map_campus(name: str) -> str | None:
    return enrollment_key_from_sis(name)


# ── Parsing ──────────────────────────────────────────────────────────

def parse_matrix() -> dict[str, list[str]]:
    with open(MATRIX_FILE) as f:
        rows = list(csv.reader(f))
    headers = rows[3]
    result: dict[str, list[str]] = {}
    for row in rows[4:]:
        campus = row[0].strip().rstrip(" ") if row[0] else ""
        if not campus or campus in ("TOTAL PROGRAMS", "Multiple Careers"):
            continue
        offered = []
        for i in range(1, min(len(row), len(headers))):
            val = row[i].strip()
            pos_name = re.sub(r"\s+", " ", headers[i].strip().replace("\n", " "))
            pos_name = re.sub(r"(\w)- ", r"\1 - ", pos_name)
            if val and val not in ("0", "") and pos_name:
                offered.append(pos_name)
        if offered:
            result[campus] = offered
    return result


def _l1_course_set() -> set[int]:
    valid: set[int] = set()
    for courses in FIRST_COURSES.values():
        valid.update(courses)
    return valid


def detect_enrollment_format(filepath: str | Path) -> str:
    """Return 'catalog' (student×course) or 'by_course' (EnrollmentByCourse)."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wb.close()
    names = {str(h).strip().lower() for h in headers if h}
    if "studentnumber" in names and "coursenumber" in names:
        return "catalog"
    return "by_course"


def parse_enrollment(filepath: str | Path) -> dict[str, dict[int, int]]:
    """EnrollmentByCourse: sum Grade-9 seat counts per campus/course."""
    valid_courses = _l1_course_set()

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active
    result: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))

    for row in ws.iter_rows(min_row=2, values_only=True):
        school, crs_num, ninth = row[1], row[9], row[26]
        if not school or not crs_num:
            continue
        try:
            crs_num = int(crs_num)
            ninth = int(ninth) if ninth else 0
        except (ValueError, TypeError):
            continue
        if ninth <= 0 or crs_num not in valid_courses:
            continue
        campus = _map_campus(school.strip())
        if campus:
            result[campus][crs_num] += ninth

    wb.close()
    return dict(result)


def parse_enrollment_unique_catalog(filepath: str | Path) -> dict[str, dict[int, int]]:
    """Student course catalog: unique Grade-9 StudentNumbers per campus/L1 course.

    Expected columns (CTE-CurrentCourseCatalog style):
    StudentNumber, FullName, GradeLevel, SchoolNumber, SchoolName, …,
    CourseNumber, CourseName, …
    """
    valid_courses = _l1_course_set()
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(h).strip() if h else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def col(*names: str) -> int | None:
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()]
        return None

    i_sn = col("StudentNumber")
    i_grade = col("GradeLevel")
    i_school = col("SchoolName")
    i_crs = col("CourseNumber")
    if None in (i_sn, i_grade, i_school, i_crs):
        wb.close()
        raise ValueError(
            "Catalog file missing required columns "
            "(StudentNumber, GradeLevel, SchoolName, CourseNumber)"
        )

    seen: dict[str, dict[int, set]] = defaultdict(lambda: defaultdict(set))
    for row in ws.iter_rows(min_row=2, values_only=True):
        grade = row[i_grade]
        if grade != 9 and str(grade).strip() != "9":
            continue
        sn = row[i_sn]
        school = row[i_school]
        crs_raw = row[i_crs]
        if not sn or not school or crs_raw is None:
            continue
        try:
            crs_num = int(crs_raw)
        except (ValueError, TypeError):
            continue
        if crs_num not in valid_courses:
            continue
        campus = _map_campus(str(school).strip())
        if campus:
            seen[campus][crs_num].add(sn)

    wb.close()
    return {
        campus: {crs: len(students) for crs, students in courses.items()}
        for campus, courses in seen.items()
    }


def resolve_enrollment(campus_courses: dict[int, int],
                       offered_pos: list[str]) -> dict[str, int]:
    course_claimants: dict[int, list[str]] = defaultdict(list)
    for pos in offered_pos:
        for crs in FIRST_COURSES.get(pos, []):
            course_claimants[crs].append(pos)

    totals: dict[str, float] = defaultdict(float)
    for crs, claimants in course_claimants.items():
        enrolled = campus_courses.get(crs, 0)
        if enrolled > 0:
            share = enrolled / len(claimants)
            for pos in claimants:
                totals[pos] += share

    return {k: round(v) for k, v in totals.items()}


SEAT_TO_MATRIX: dict[str, str] = {
    "automotive and collision repair": "Automotive Collision",
    "automotive": "Automotive Maintenance",
    "graphic design": "Graphic Design",
    "graphic design and interactive media": "Graphic Design",
    "engineering": "Engineering Foundations",
    "engineering foundations": "Engineering Foundations",
    "culinary arts": "Culinary Arts",
    "culinary": "Culinary Arts",
    "marketing": "Marketing & Sales",
    "marketing and sales": "Marketing & Sales",
    "biomedical": "Biomedical Science",
    "biomedical science": "Biomedical Science",
    "entrepreneurship": "Entrepreneurship",
    "accounting": "Accounting & Financial Services",
    "accounting and financial services": "Accounting & Financial Services",
    "acounting and financial services": "Accounting & Financial Services",  # Notes typo
    "business management": "Business Management",
    "entreprenuership": "Entrepreneurship",  # Notes typo
    "law enforcement": "Law Enforcement",
    "digital communications": "Digital Communications",
    "fire science": "Emergency Services - Firefighter",
    "animal science": "Animal Science",
    "agricultural technology": "Applied Agricultural Engineering - Welding",
    "agricultural technology and mechanical systems": "Applied Agricultural Engineering - Welding",
    "information technology support and services": "Information Technology Support & Services",
    "it support": "Information Technology Support & Services",
    "web development": "Web Development",
    "legal studies": "Legal Studies",
    "teaching and training": "Teaching & Training",
    "cosmetology": "Cosmetology & Personal Care",
    "cosmetology and personal care services": "Cosmetology & Personal Care",
    "lodging and resort management": "Lodging & Resort Management",
    "lodging": "Lodging & Resort Management",
    "architectural drafting and design": "Architectural Design",
    "architectural design": "Architectural Design",
    "construction/ carpentry": "Carpentry",
    "construction/carpentry": "Carpentry",
    "carpentry": "Carpentry",
    "hvac (heating, ventilation, and air conditioning)": "HVAC & Sheet Metal",
    "hvac and sheet metal": "HVAC & Sheet Metal",
    "hvac": "HVAC & Sheet Metal",
    "electrical & solar": "Electrical",
    "electrical and solar": "Electrical",
    "electrical": "Electrical",
    "plumbing": "Plumbing & Pipe Fitting",
    "plumbing and pipefitting": "Plumbing & Pipe Fitting",
    "cybersecurity": "Cybersecurity",
    "software development & game design": "Programming & Software Development",
    "software development and game design": "Programming & Software Development",
    "programming": "Programming & Software Development",
    "programming and software development": "Programming & Software Development",
    "electronics technology": "Electronics Technology",
    "electronics": "Electronics Technology",
    "welding": "Welding",
    "aviation (pilots)": "Aviation Flight",
    "aviation": "Aviation Flight",
    "aviation maintenance": "Aviation Maintenance",
    "interior design": "Interior Design",
    "plant science": "Horticulture Science",
    "fashion design": "Fashion Design",
    "diagnostic & therapeutic services (patient care technician)":
        "Healthcare Therapeutic - Patient Care Tech",
    "diagnostic and therapeutic services (patient care technician)":
        "Healthcare Therapeutic - Patient Care Tech",
    "diagnostic & therapeutic services (ekg/ phelobotomy)":
        "Healthcare Diagnostic - ECG/EKG & Phlebotomy",
    "diagnostic & therapeutic services (ekg/ phlebotomy)":
        "Healthcare Diagnostic - ECG/EKG & Phlebotomy",
    "diagnostic and therapeutic services (ekg/ phelobotomy)":
        "Healthcare Diagnostic - ECG/EKG & Phlebotomy",
    "diagnostic and therapeutic services (ekg/ phlebotomy)":
        "Healthcare Diagnostic - ECG/EKG & Phlebotomy",
    "diagnostic and therapeutic services": "Healthcare Therapeutic - Patient Care Tech",
    "dental assistant": "Healthcare Therapeutic - Registered Dental Asst.",
    "emt (emergency medical technician)": "Healthcare Therapeutic - Emergency Med. Tech.",
    "emt": "Healthcare Therapeutic - Emergency Med. Tech.",
    # PoS Notes (Internal) aliases
    "barbery": "Cosmetology & Personal Care",
    "barbering": "Cosmetology & Personal Care",
    "automotive - collision magnet": "Automotive Collision",
    "automotive - tech magnet": "Automotive Maintenance",
    "automotive tech magnet": "Automotive Maintenance",
    "aviation maintance - aviation magnet": "Aviation Maintenance",
    "aviation maintenance - aviation magnet": "Aviation Maintenance",
    "architectural design - architecture magnet": "Architectural Design",
    "interior design magnet": "Interior Design",
    "fashion design magnet": "Fashion Design",
    "hvac magnet": "HVAC & Sheet Metal",
    "horticulture": "Horticulture Science",
    "floral design": "Floral Design",
    "legal studies - adv. social sciences magnet": "Legal Studies",
    "princ law & pub safety": "Law Enforcement",
    "principles of law, public service, corr., security": "Law Enforcement",
    "principles of law, public safety, corrections, and security": "Law Enforcement",
    "principles of business, marketing, and finance": "Business Management",
    "principles of business marketing and finance": "Business Management",
    "principles of business marketing & finance (afs 1)": "Accounting & Financial Services",
    "principles of business, marketing, and finance (afs 1)": "Accounting & Financial Services",
    "principles of business, marketing, and finance (entr 1)": "Entrepreneurship",
    "principles of business, marketing, and finance (m&s 1)": "Marketing & Sales",
    "principles of business, marketing, and finance (bmgt 1)": "Business Management",
    "diagnostic and therapeutic services (medical laboratory)":
        "Healthcare Diagnostic - ECG/EKG & Phlebotomy",
    "diagnostic and therapeutic services (clinical medical assisting)":
        "Healthcare Therapeutic - Patient Care Tech",
    "diagnostic and therapeutic services (dental assisting)":
        "Healthcare Therapeutic - Registered Dental Asst.",
    "diagnostic and therapeutic services (patient care tech)":
        "Healthcare Therapeutic - Patient Care Tech",
    "diagnostic and therapeutic services (exercise science)":
        "Exercise Science & Wellness",
    "advanced manufacturing engineering foundations": "Engineering Foundations",
    "advanced manufacturing  engineering foundations": "Engineering Foundations",
    "family and community services": "Family & Community Services",
    "princiapls of agriculture, food and natural resources":
        "Environmental & Natural Resources",
    "principles of agriculture, food and natural resources":
        "Environmental & Natural Resources",
    "graphic design and multimedia arts": "Graphic Design",
    "graphic design & multimedia arts": "Graphic Design",
    "graphic design magnet": "Graphic Design",
    "digital communications magnet": "Digital Communications",
    "cosmetology and personal care services magnet": "Cosmetology & Personal Care",
    "family & community services": "Family & Community Services",
    "family and community services": "Family & Community Services",
    "human services": "Family & Community Services",
    "graphic design & multimedia arts - (animation)": "Animation",
    "graphic design and multimedia arts - (animation)": "Animation",
    "graphic design & multimedia arts - (video gaming)": "Video Game Design",
    "graphic design and multimedia arts - (video gaming)": "Video Game Design",
    "animation": "Animation",
    "video game design": "Video Game Design",
    "video gaming": "Video Game Design",
    "digital media": "Digital Communications",
    "architecture": "Architectural Design",
    "agricuture": "Environmental & Natural Resources",
    "agriculture": "Environmental & Natural Resources",
}

CAMPUS_NORM = {
    "White": "WT White", "WT White": "WT White",
    "W. Hutchins": "Wilmer-Hutchins", "W Hutchins": "Wilmer-Hutchins",
    "Wilson": "Woodrow Wilson", "Pinkston": "Pinkston",
    "L.G. Pinkston": "Pinkston", "Woodrow Wilson": "Woodrow Wilson",
    "Wilmer-Hutchins": "Wilmer-Hutchins",
}

CI_TITLE_MAP = {
    "career institute north": "Career Institute - North",
    "career institute south": "Career Institute - South",
    "career institute east": "Career Institute - East",
}


def norm_pos(raw: str) -> str | None:
    cleaned = re.sub(r"\s+", " ", raw.lower().strip().replace("\n", " "))
    cleaned = re.sub(r"\s*program\s*of\s*study.*$", "", cleaned).strip()
    cleaned = cleaned.replace("phelobotomy", "phlebotomy")
    cleaned = cleaned.replace("acounting", "accounting")
    cleaned = cleaned.replace("entreprenuership", "entrepreneurship")
    # Drop status suffixes from Notes workbook
    cleaned = re.sub(
        r"\s*[\(\-]?\s*(phasing\s*(out|in)|sunsetting|closed|program is phasing"
        r"[^)]*)\)?\s*$",
        "",
        cleaned,
        flags=re.I,
    ).strip(" -")
    # Also strip trailing "sunsetting YYYY-YYYY" without parens
    cleaned = re.sub(
        r"\s*\(?(?:sunsetting|phasing\s*out|closed)[^)]*\)?\s*$",
        "",
        cleaned,
        flags=re.I,
    ).strip(" -")
    if cleaned in SEAT_TO_MATRIX:
        return SEAT_TO_MATRIX[cleaned]
    no_paren = re.sub(r"\s*\(.*$", "", cleaned).strip()
    if no_paren in SEAT_TO_MATRIX:
        return SEAT_TO_MATRIX[no_paren]
    # Exact FIRST_COURSES key match
    for key in FIRST_COURSES:
        if key.lower() == cleaned or key.lower() == no_paren:
            return key
    for pattern, canonical in sorted(SEAT_TO_MATRIX.items(), key=lambda x: -len(x[0])):
        if pattern in cleaned or cleaned in pattern:
            return canonical
    for key in FIRST_COURSES:
        kl = key.lower()
        if kl in cleaned or cleaned in kl:
            return key
    return None


def parse_pos_notes_seats() -> dict[str, dict[str, float]]:
    """L1 Max Capacity from 25-26 Programs Of Study Notes (Internal).xlsx.

    Used only to fill blank seat goals — never overwrites the seat workbook.
    """
    if not POS_NOTES_FILE.is_file():
        return {}

    wb = openpyxl.load_workbook(str(POS_NOTES_FILE), data_only=True)
    result: dict[str, dict[str, float]] = defaultdict(dict)

    for sheet_name, campus in NOTES_SHEET_TO_CAMPUS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        current_pos: str | None = None
        in_table = False
        level_col = 4
        cap_col = 6
        l1_caps: list[float] = []
        l2_caps: list[float] = []
        l3_caps: list[float] = []

        def flush():
            nonlocal l1_caps, l2_caps, l3_caps, current_pos
            if current_pos:
                total = sum(l1_caps)
                # Some Notes sheets leave L1 max at 0 — fall back to L2, then L3
                # (sunsetting pathways often only list upper-level courses)
                if total <= 0:
                    total = sum(l2_caps)
                if total <= 0:
                    total = sum(l3_caps)
                if total > 0:
                    prev = result[campus].get(current_pos, 0)
                    result[campus][current_pos] = max(prev, total)
            l1_caps = []
            l2_caps = []
            l3_caps = []

        def _parse_cap(v) -> float | None:
            try:
                f = float(v)  # type: ignore[arg-type]
            except (TypeError, ValueError):
                return None
            return f if f > 0 else None

        def row_level_cap(r: int) -> tuple[str | None, float | None]:
            """Read Course Level + Max Capacity, tolerating shifted Notes columns."""
            level = ws.cell(r, level_col).value
            lvl = str(level).lower() if level is not None else ""
            if "level" in lvl:
                cap_f = _parse_cap(ws.cell(r, cap_col).value)
                if cap_f is not None:
                    return lvl, cap_f
            # Scan row when Checked/Level/Credits columns are misaligned (e.g. Pinkston)
            level_at = None
            for c in range(2, 12):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "level" in v.lower():
                    level_at = c
                    lvl = v.lower()
                    break
            if level_at is None:
                return None, None
            # Prefer two cols after Level (skip credits). Credit-sized values
            # in the +1 slot are almost always "# of Credits", not capacity.
            creditish = {0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0}
            cap_f = None
            for off in (2, 3, 1):
                cand = _parse_cap(ws.cell(r, level_at + off).value)
                if cand is None:
                    continue
                if off == 1 and cand in creditish:
                    continue
                cap_f = cand
                break
            if cap_f is None:
                cap_f = _parse_cap(ws.cell(r, cap_col).value)
            return lvl, cap_f

        def looks_like_pos_header(text: str) -> str | None:
            """Notes sometimes omit the Priority line and jump straight to PoS name."""
            low = text.lower().strip()
            if not low or low == "course title" or low.startswith("campus:"):
                return None
            if "priority" in low or low.startswith("total"):
                return None
            if low.startswith("principal") or low.startswith("executive"):
                return None
            if "elective" in low or "science cte" in low:
                return None
            # Skip closed pathways (no useful L1 capacity)
            if "closed" in low:
                return None
            return norm_pos(text)

        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if not isinstance(a, str):
                continue
            a_st = a.strip()
            a_low = a_st.lower()

            if "program of study priority" in a_low:
                flush()
                in_table = False
                after = a_st.split(":", 1)[1].strip() if ":" in a_st else ""
                current_pos = norm_pos(after) if after else None
                if current_pos is None:
                    for rr in range(r + 1, min(r + 4, ws.max_row + 1)):
                        v = ws.cell(rr, 1).value
                        if isinstance(v, str) and v.strip():
                            if v.strip().lower() == "course title":
                                break
                            current_pos = looks_like_pos_header(v.strip()) or norm_pos(v.strip())
                            if current_pos:
                                break
                continue

            if a_low == "course title":
                in_table = True
                level_col, cap_col = 4, 6
                for c in range(1, 12):
                    h = ws.cell(r, c).value
                    if not isinstance(h, str):
                        continue
                    hl = h.lower()
                    if "course level" in hl:
                        level_col = c
                    if "max capacity" in hl or "estimated # of students" in hl:
                        cap_col = c
                continue

            # Totals may sit in col A or col D depending on the campus sheet
            is_totals = a_low.startswith("total")
            if not is_totals:
                for c in range(2, 6):
                    tv = ws.cell(r, c).value
                    if isinstance(tv, str) and tv.strip().lower().startswith("total"):
                        is_totals = True
                        break
            if in_table and is_totals:
                flush()
                in_table = False
                continue

            # Bare PoS name row (Notes often omit Priority header — e.g. Skyline)
            maybe = looks_like_pos_header(a_st)
            if maybe and a_low != "course title":
                ahead = False
                for rr in range(r + 1, min(r + 3, ws.max_row + 1)):
                    v = ws.cell(rr, 1).value
                    if isinstance(v, str) and v.strip().lower() == "course title":
                        ahead = True
                        break
                if ahead:
                    flush()
                    in_table = False
                    current_pos = maybe
                    continue

            if in_table and current_pos and a_st:
                lvl, cap_f = row_level_cap(r)
                if not lvl or cap_f is None:
                    continue
                if "1" in lvl and "2" not in lvl and "3" not in lvl and "4" not in lvl:
                    l1_caps.append(cap_f)
                elif "2" in lvl and "1" not in lvl and "3" not in lvl and "4" not in lvl:
                    l2_caps.append(cap_f)
                elif "3" in lvl and "1" not in lvl and "2" not in lvl and "4" not in lvl:
                    l3_caps.append(cap_f)

        flush()

    wb.close()
    return dict(result)


def parse_ci_feeder_allocations() -> dict[str, dict[str, dict[str, float]]]:
    wb = openpyxl.load_workbook(str(SEATS_FILE), data_only=True)
    ws = wb["FRESHMAN 26-27 Career Institute"]
    result: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(dict)
    )
    current_ci: str | None = None
    col_headers: dict[int, str] = {}

    for row in range(1, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value

        if isinstance(a, str) and "career institute" in a.lower():
            key = re.sub(r"\s*\(.*$", "", a.lower()).strip()
            current_ci = None
            for pattern, name in CI_TITLE_MAP.items():
                if pattern in key:
                    current_ci = name
                    break
            col_headers = {}
            continue

        draft = isinstance(b, str) and "draft" in b.lower()
        has_pos_header = any(
            isinstance(ws.cell(row, c).value, str)
            and "program of study" in str(ws.cell(row, c).value).lower()
            for c in range(5, 22)
        )
        if draft or has_pos_header:
            col_headers = {}
            for c in range(5, 22):
                raw = ws.cell(row, c).value
                if isinstance(raw, str) and "program of study" in raw.lower():
                    pos = norm_pos(raw)
                    if pos:
                        col_headers[c] = pos
            continue

        if not current_ci or not isinstance(b, str):
            continue
        campus_raw = b.strip()
        if (not campus_raw or campus_raw.upper().startswith("TOTAL")
                or "ALLOWED" in campus_raw.upper()
                or campus_raw in ("Comprehensive HS", "DRAFT NUMBERS!", "CAMPUS")):
            continue

        campus = CAMPUS_NORM.get(campus_raw.strip(), campus_raw.strip())
        for c, pos in col_headers.items():
            val = ws.cell(row, c).value
            try:
                seats = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                seats = 0.0
            if seats > 0:
                result[current_ci][campus][pos] = seats

    wb.close()
    return {ci: dict(feeders) for ci, feeders in result.items()}


def parse_seats() -> tuple[
    dict[str, dict[str, float]],
    dict[str, dict[str, dict[str, float]]],
]:
    result: dict[str, dict[str, float]] = defaultdict(dict)
    wb = openpyxl.load_workbook(str(SEATS_FILE), data_only=True)
    ws = wb["Updated FRESHMAN 26-27 CTE Comp"]
    skip = {"CAMPUS", "", "TOTALS"}
    row = 3
    while row <= ws.max_row:
        raw = ws.cell(row, 2).value
        if not raw or not isinstance(raw, str) or raw.strip() in skip:
            row += 1
            continue
        campus = CAMPUS_NORM.get(raw.strip(), raw.strip())
        for col in range(15, 33):
            pos_raw = ws.cell(row, col).value
            seats_val = ws.cell(row + 1, col).value if row + 1 <= ws.max_row else None
            if pos_raw and isinstance(pos_raw, str):
                pos_name = norm_pos(pos_raw)
                seats = float(seats_val) if seats_val else 0
                if pos_name and seats > 0:
                    result[campus][pos_name] = seats
        row += 2
    wb.close()

    ci_feeders = parse_ci_feeder_allocations()
    for _ci, feeders in ci_feeders.items():
        for campus, pos_seats in feeders.items():
            for pos, seats in pos_seats.items():
                result[campus][pos] = seats

    # Fill blanks from PoS Notes L1 max capacity (do not overwrite known goals).
    # When Notes uses a different PoS name than the course matrix (e.g. Notes
    # "Web Development" vs matrix "IT Support"), map onto the blank matrix
    # program that shares the same L1 course(s).
    notes = parse_pos_notes_seats()
    matrix = parse_matrix()
    for campus, pos_seats in notes.items():
        offered = set(matrix.get(campus, []))
        for pos, seats_n in pos_seats.items():
            if seats_n <= 0:
                continue
            if pos in offered:
                if result[campus].get(pos, 0) <= 0:
                    result[campus][pos] = seats_n
                continue
            # Notes name not on matrix — prefer blank siblings sharing L1 courses
            notes_crs = set(FIRST_COURSES.get(pos, []))
            mapped = False
            for mpos in offered:
                if result[campus].get(mpos, 0) > 0:
                    continue
                if notes_crs & set(FIRST_COURSES.get(mpos, [])):
                    result[campus][mpos] = seats_n
                    mapped = True
            if not mapped and result[campus].get(pos, 0) <= 0:
                result[campus][pos] = seats_n
            # Also fill blank matrix siblings even when Notes name already has seats
            # elsewhere (e.g. Digital Comms seats + Graphic Design on matrix)
            for mpos in offered:
                if mpos == pos:
                    continue
                if result[campus].get(mpos, 0) > 0:
                    continue
                if notes_crs & set(FIRST_COURSES.get(mpos, [])):
                    result[campus][mpos] = seats_n

    return dict(result), ci_feeders


def _program_rows(
    pos_list: list[str],
    enrolled_map: dict[str, int],
    seats_map: dict[str, float],
    *,
    ci_pos_for_campus: set[str] | None = None,
) -> list[dict]:
    """Build program rows. CI-feeder flag only when this campus has CI seat goals."""
    programs = []
    ci_set = ci_pos_for_campus or set()
    for pos in pos_list:
        crs_list = FIRST_COURSES.get(pos, [])
        seat_count = seats_map.get(pos)
        enrolled = enrolled_map.get(pos, 0)
        seats_i = int(seat_count) if seat_count and seat_count > 0 else None
        # No seats and no students — omit (avoids blank CI feeder rows)
        if seats_i is None and not enrolled:
            continue
        pct = (enrolled / seats_i) if seats_i else None
        programs.append({
            "name": pos,
            "courses": crs_list,
            "seats": seats_i,
            "enrolled": enrolled,
            "pct": round(pct, 3) if pct is not None else None,
            # Only true CI feeders (seat allocation), never name-only CI_PROGRAMS
            "is_ci_program": pos in ci_set and seats_i is not None,
        })
    return programs


# ── Main pipeline ────────────────────────────────────────────────────

def process_enrollment_file(
    filepath: str | Path,
    date_label: str,
    *,
    unique_students: bool | None = None,
) -> list[dict]:
    """Run the full pipeline and return campus data rows for storage.

    unique_students:
      None  — auto-detect from file headers (catalog vs EnrollmentByCourse)
      True  — count unique Grade-9 students per campus/L1 course
      False — sum EnrollmentByCourse seat rows (legacy)
    """
    matrix = parse_matrix()
    if unique_students is None:
        unique_students = detect_enrollment_format(filepath) == "catalog"
    enrollment = (
        parse_enrollment_unique_catalog(filepath)
        if unique_students
        else parse_enrollment(filepath)
    )
    seats, ci_feeders = parse_seats()

    # Campus → PoS that have real CI feeder seat goals (not Notes fill-ins)
    ci_pos_by_campus: dict[str, set[str]] = defaultdict(set)
    for _ci, feeders in ci_feeders.items():
        for feeder, pos_seats in feeders.items():
            for pos, n in pos_seats.items():
                if n and n > 0:
                    ci_pos_by_campus[feeder].add(pos)

    campus_rows: list[dict] = []

    # Comprehensive + magnet campuses (exclude bare CI matrix rows — rebuilt below)
    for campus in sorted(matrix.keys()):
        if campus.startswith("Career Institute"):
            continue
        seats_map = seats.get(campus, {})
        pos_list = list(matrix[campus])
        for pos in seats_map:
            if pos not in pos_list:
                pos_list.append(pos)
        courses = enrollment.get(campus, {})
        enrolled_map = resolve_enrollment(courses, pos_list)
        programs = _program_rows(
            pos_list,
            enrolled_map,
            seats_map,
            ci_pos_for_campus=ci_pos_by_campus.get(campus, set()),
        )
        campus_rows.append({
            "campus": campus,
            "is_ci": False,
            "total_seats": sum(p["seats"] or 0 for p in programs),
            "total_enrolled": sum(p["enrolled"] for p in programs),
            "programs": programs,
        })

    # CI regions with feeder high school breakdowns
    for ci in sorted(ci_feeders.keys()):
        feeders_out = []
        ci_seats = 0
        ci_enrolled = 0
        for feeder in sorted(ci_feeders[ci].keys()):
            seats_map = ci_feeders[ci][feeder]
            # Skip feeders with no seat goals
            if not seats_map or not any(v > 0 for v in seats_map.values()):
                continue
            pos_list = sorted(p for p, n in seats_map.items() if n and n > 0)
            if not pos_list:
                continue
            courses = enrollment.get(feeder, {})
            enrolled_map = resolve_enrollment(courses, pos_list)
            programs = _program_rows(
                pos_list,
                enrolled_map,
                seats_map,
                ci_pos_for_campus=set(pos_list),
            )
            if not programs:
                continue
            f_seats = sum(p["seats"] or 0 for p in programs)
            f_enrolled = sum(p["enrolled"] for p in programs)
            feeders_out.append({
                "campus": feeder,
                "total_seats": f_seats,
                "total_enrolled": f_enrolled,
                "programs": programs,
            })
            ci_seats += f_seats
            ci_enrolled += f_enrolled

        campus_rows.append({
            "campus": ci,
            "is_ci": True,
            "total_seats": ci_seats,
            "total_enrolled": ci_enrolled,
            "programs": [],
            "feeders": feeders_out,
        })

    return campus_rows
