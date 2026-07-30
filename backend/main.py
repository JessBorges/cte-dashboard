"""FastAPI backend for CTE Data Dashboard (Enrollment + IBC)."""

from __future__ import annotations

import importlib
import os
import sys
import shutil
import zipfile
import tempfile
from pathlib import Path

from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse

from database import init_db, save_snapshot, list_snapshots, get_snapshot, get_latest_snapshot, delete_snapshot
from processor import process_enrollment_file, detect_enrollment_format
from campus_norm import campus_key
import ibc_bridge

UPLOAD_DIR = Path(__file__).resolve().parent.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

IBC_DATA_DIR = Path(__file__).resolve().parent / "ibc_data"
IBC_DATA_DIR.mkdir(exist_ok=True)

TRACKER_DIR = Path(__file__).resolve().parent.parent / "uploads" / "tracker"
TRACKER_DIR.mkdir(parents=True, exist_ok=True)

FRONTEND_DIST = Path(__file__).resolve().parent.parent / "frontend" / "dist"

ALLOWED_ORIGINS = os.environ.get(
    "CORS_ORIGINS",
    "http://localhost:5173,http://127.0.0.1:5173",
).split(",")

app = FastAPI(title="CTE Data Dashboard API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in ALLOWED_ORIGINS],
    allow_methods=["*"],
    allow_headers=["*"],
)

init_db()


@app.get("/api/snapshots")
def api_list_snapshots():
    return list_snapshots()


@app.get("/api/snapshots/latest")
def api_latest_snapshot():
    data = get_latest_snapshot()
    if not data:
        raise HTTPException(404, "No snapshots yet. Upload enrollment data first.")
    for c in data.get("campuses", []):
        c["campus_key"] = campus_key(c.get("campus", ""))
        c["display_name"] = c.get("campus", "")
    return data


@app.get("/api/snapshots/{snap_id}")
def api_get_snapshot(snap_id: int):
    data = get_snapshot(snap_id)
    if not data:
        raise HTTPException(404, "Snapshot not found")
    for c in data.get("campuses", []):
        c["campus_key"] = campus_key(c.get("campus", ""))
        c["display_name"] = c.get("campus", "")
    return data


@app.post("/api/upload")
async def api_upload(
    file: UploadFile = File(...),
    label: str = Form(""),
    date_label: str = Form("7/27"),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Please upload an .xlsx file")

    dest = UPLOAD_DIR / file.filename
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)

    fmt = detect_enrollment_format(str(dest))
    unique = fmt == "catalog"
    if not label:
        label = (
            f"Week of {date_label} (unique G9)"
            if unique
            else f"Week of {date_label}"
        )

    try:
        campus_rows = process_enrollment_file(
            str(dest), date_label, unique_students=unique
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"Failed to process file: {exc}") from exc

    snap_id = save_snapshot(label, date_label, file.filename, campus_rows)

    return {
        "snapshot_id": snap_id,
        "label": label,
        "campuses": len(campus_rows),
        "source_format": fmt,
        "unique_students": unique,
    }


@app.delete("/api/snapshots/{snap_id}")
def api_delete_snapshot(snap_id: int):
    delete_snapshot(snap_id)
    return {"ok": True}


@app.get("/api/ibc/status")
def api_ibc_status():
    return ibc_bridge.status()


@app.post("/api/ibc/refresh")
def api_ibc_refresh():
    out = ibc_bridge.refresh_mapping(force=True)
    if not out.get("ok"):
        raise HTTPException(503, out.get("error") or "IBC refresh failed")
    return out


@app.get("/api/ibc/summary")
def api_ibc_summary():
    return _ibc_or_503(ibc_bridge.summary)


@app.get("/api/ibc/campuses")
def api_ibc_campuses():
    return _ibc_or_503(ibc_bridge.campuses_list)


@app.get("/api/ibc/campuses/{key}")
def api_ibc_campus_detail(key: str):
    def _load():
        detail = ibc_bridge.campus_detail(key)
        if not detail:
            raise HTTPException(404, f"Campus not found: {key}")
        return detail
    return _ibc_or_503(_load)


@app.get("/api/ibc/pos")
def api_ibc_pos():
    return _ibc_or_503(ibc_bridge.pos_list)


@app.get("/api/ibc/certs")
def api_ibc_certs():
    return _ibc_or_503(ibc_bridge.certs_list)


@app.get("/api/campus/{key}/combined")
def api_campus_combined(key: str):
    enroll = get_latest_snapshot()
    enroll_campus = None
    if enroll:
        for c in enroll.get("campuses", []):
            if campus_key(c.get("campus", "")) == campus_key(key) or c.get("campus", "").lower() == key.lower():
                enroll_campus = c
                break

    ibc = None
    ibc_error = None
    try:
        ibc = ibc_bridge.campus_detail(key)
    except Exception as exc:
        ibc_error = str(exc)

    if not enroll_campus and not ibc:
        raise HTTPException(404, f"No enrollment or IBC data for campus: {key}")

    return {
        "campus_key": campus_key(key),
        "enrollment": enroll_campus,
        "enrollment_snapshot": enroll["snapshot"] if enroll else None,
        "ibc": ibc,
        "ibc_error": ibc_error,
    }


@app.post("/api/ibc/export-portfolio")
def api_ibc_export():
    return _ibc_or_503(ibc_bridge.export_portfolio)


@app.post("/api/ibc/upload-data")
async def api_ibc_upload_data(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith(".zip"):
        raise HTTPException(400, "Please upload a .zip file of your IBC Tiers data")
    zip_path = IBC_DATA_DIR / file.filename
    with open(zip_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(IBC_DATA_DIR)
    except zipfile.BadZipFile:
        zip_path.unlink(missing_ok=True)
        raise HTTPException(400, "Invalid zip file")
    zip_path.unlink(missing_ok=True)
    nested = IBC_DATA_DIR / "IBC Tiers"
    data_dir = str(nested) if nested.is_dir() else str(IBC_DATA_DIR)
    os.environ["IBC_DATA_DIR"] = data_dir
    for _mod in list(sys.modules):
        if _mod == "ibc_weekly" or _mod.startswith("build_tier1_pos_mapping"):
            sys.modules.pop(_mod, None)
    importlib.reload(ibc_bridge)
    out = ibc_bridge.refresh_mapping(force=True)
    return {
        "ok": out.get("ok", False),
        "error": out.get("error"),
        "path": data_dir,
    }


@app.post("/api/ibc/upload-tracker")
async def api_ibc_upload_tracker(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Please upload an .xlsx workbook")
    dest = TRACKER_DIR / file.filename
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)
    os.environ["IBC_TRACKER_PATH"] = str(dest)
    for _mod in list(sys.modules):
        if _mod == "ibc_weekly" or _mod.startswith("build_tier1_pos_mapping"):
            sys.modules.pop(_mod, None)
    importlib.reload(ibc_bridge)
    out = ibc_bridge.refresh_mapping(force=True)
    return {
        "ok": out.get("ok", False),
        "error": out.get("error"),
        "file": str(dest.name),
    }


def _ibc_or_503(fn):
    try:
        return fn()
    except Exception as exc:
        raise HTTPException(
            503,
            f"IBC data unavailable: {exc}. "
            "Ensure Public/IBC Tiers sources exist, then POST /api/ibc/refresh.",
        ) from exc


@app.get("/api/export/xlsx")
def api_export_xlsx():
    """Multi-sheet XLSX export of all dashboard data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from campus_norm import campus_key as ck

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def style(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border
        ws.freeze_panes = "A2"

    # --- Sheet 1: Enrollment ---
    ws1 = wb.active
    ws1.title = "Enrollment"
    h1 = ["Campus", "Campus Key", "Type", "Total Seats", "Total Enrolled", "Fill %",
          "Programs (count)"]
    style(ws1, h1)
    enroll = get_latest_snapshot()
    campuses = (enroll or {}).get("campuses", [])
    for i, c in enumerate(campuses, 2):
        seats = c.get("total_seats", 0) or 0
        enrolled = c.get("total_enrolled", 0) or 0
        fill = round(100 * enrolled / seats, 1) if seats else 0
        ws1.append([
            c.get("campus", ""), ck(c.get("campus", "")),
            "CI" if c.get("is_ci") else "Comprehensive",
            seats, enrolled, fill,
            len(c.get("programs", [])),
        ])

    # --- Sheet 2: IBC Summary ---
    ws2 = wb.create_sheet("IBC Summary")
    try:
        s = ibc_bridge.summary()
        wk = s.get("weekly") or {}
        ws2.append(["Metric", "Value"])
        style(ws2, ["Metric", "Value"])
        rows = [
            ("Built At", s.get("built_at")),
            ("All Earns (3yr)", s.get("all_earns_3yr")),
            ("Tier 1 (3yr)", s.get("tier1_3yr")),
            ("POS Offered", s.get("pos_offered")),
            ("POS T1 Eligible", s.get("pos_t1_eligible")),
            ("Weekly Tracker Loaded", "Yes" if wk.get("title") else "No"),
            ("Weekly Earned", wk.get("earned")),
            ("Weekly Projected", wk.get("projected")),
            ("Weekly T1", wk.get("t1")),
            ("Weekly T2", wk.get("t2")),
            ("Weekly T3", wk.get("t3")),
            ("Weekly Cert Count", wk.get("cert_count")),
            ("Weekly Campus Count", wk.get("campus_count")),
        ]
        if wk.get("certs_by_tier"):
            ct = wk["certs_by_tier"]
            ws2.append([])
            ws2.append(["--- Certs by Tier ---"])
            ws2.append(["Tier", "Cert Name", "Earned", "Projected", "G12"])
            for tk in ["t1", "t2", "t3", "none"]:
                for cr in ct.get(tk, []):
                    ws2.append([tk.upper(), cr["name"], cr["earned"],
                                cr.get("projected", 0), cr.get("g12", 0)])
        cp = s.get("completer", {})
        ws2.append([])
        ws2.append(["--- Completer Stats ---"])
        for k, v in cp.items():
            if k != "method":
                ws2.append([k, v])
    except Exception:
        ws2.append(["Error", "IBC data not available"])

    # --- Sheet 3: IBC Campuses ---
    ws3 = wb.create_sheet("IBC Campuses")
    h3 = ["Campus Key", "Display Name", "Attempts", "Earned", "Pass Rate %",
          "T1 Earned", "T2 Earned", "T3 Earned", "None Earned", "T1 Attempts"]
    style(ws3, h3)
    try:
        for row in ibc_bridge.campuses_list():
            ws3.append([
                row.get("campus_key"), row.get("display_name"),
                row.get("attempts", 0), row.get("earned", 0),
                row.get("pass_rate", 0),
                row.get("t1_earned", 0), row.get("t2_earned", 0),
                row.get("t3_earned", 0), row.get("none_earned", 0),
                row.get("t1_attempts", 0),
            ])
    except Exception:
        ws3.append(["IBC data not available"])

    # --- Sheet 4: Programs of Study ---
    ws4 = wb.create_sheet("Programs of Study")
    h4 = ["POS Code", "POS Name", "Program Name", "Tier 1 Eligible", "Certs"]
    style(ws4, h4)
    try:
        for pos in ibc_bridge.pos_list():
            ws4.append([
                pos.get("code"), pos.get("name"),
                pos.get("program_name", ""),
                "Yes" if pos.get("tier1_eligible") else "No",
                "; ".join(pos.get("certs", [])),
            ])
    except Exception:
        ws4.append(["IBC data not available"])

    # --- Sheet 5: Certifications ---
    ws5 = wb.create_sheet("Certifications")
    h5 = ["Code", "Name", "Tier", "Tier Key"]
    style(ws5, h5)
    try:
        for cert in ibc_bridge.certs_list():
            ws5.append([
                cert.get("code"), cert.get("name"),
                cert.get("tier", ""), cert.get("tier_key", ""),
            ])
    except Exception:
        ws5.append(["IBC data not available"])

    # --- Sheet 6: Combined Campus ---
    ws6 = wb.create_sheet("Combined Campus")
    h6 = ["Campus", "Campus Key", "Type", "Seats", "Enrolled", "Fill %",
          "IBC Attempts", "IBC Earned", "IBC Pass Rate",
          "T1 Earned", "T2 Earned", "T3 Earned"]
    style(ws6, h6)
    ibc_map = {}
    try:
        for row in ibc_bridge.campuses_list():
            ibc_map[row["campus_key"]] = row
    except Exception:
        pass
    seen = set()
    for c in campuses:
        key = ck(c.get("campus", ""))
        seen.add(key)
        ibc = ibc_map.get(key, {})
        seats = c.get("total_seats", 0) or 0
        enrolled = c.get("total_enrolled", 0) or 0
        fill = round(100 * enrolled / seats, 1) if seats else 0
        ws6.append([
            c.get("campus", ""), key,
            "CI" if c.get("is_ci") else "Comprehensive",
            seats, enrolled, fill,
            ibc.get("attempts", 0), ibc.get("earned", 0), ibc.get("pass_rate", 0),
            ibc.get("t1_earned", 0), ibc.get("t2_earned", 0), ibc.get("t3_earned", 0),
        ])
    for key, ibc in ibc_map.items():
        if key not in seen:
            ws6.append([
                key, key, "IBC only",
                0, 0, 0,
                ibc.get("attempts", 0), ibc.get("earned", 0), ibc.get("pass_rate", 0),
                ibc.get("t1_earned", 0), ibc.get("t2_earned", 0), ibc.get("t3_earned", 0),
            ])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = str(cell.value or "")
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cte_dashboard_export.xlsx"},
    )


@app.get("/api/export/csv")
def api_export_csv():
    enroll = get_latest_snapshot()
    import csv
    import io
    from campus_norm import campus_key as ck

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Campus", "Campus Key", "Type", "Total Seats", "Total Enrolled",
                 "Fill %", "IBC Attempts", "IBC Earned", "Tier 1", "Tier 2", "Tier 3"])

    campuses = (enroll or {}).get("campuses", [])
    if not campuses and not ibc_bridge.get_result():
        raise HTTPException(404, "No enrollment or IBC data to export")

    ibc_campuses = {}
    try:
        result = ibc_bridge.get_result()
        from ibc_bridge import _agg_by_campus
        ibc_campuses = _agg_by_campus(result["attempt_rows"], {})
    except Exception:
        pass

    seen = set()
    for c in campuses:
        key = ck(c.get("campus", ""))
        seen.add(key)
        ibc = ibc_campuses.get(key, {})
        seats = c.get("total_seats", 0) or 0
        enrolled = c.get("total_enrolled", 0) or 0
        fill = round(100 * enrolled / seats, 1) if seats else 0
        w.writerow([
            c.get("campus", ""), key,
            "CI" if c.get("is_ci") else "Comprehensive",
            seats, enrolled, fill,
            ibc.get("attempts", 0), ibc.get("earned", 0),
            ibc.get("t1_earned", 0), ibc.get("t2_earned", 0), ibc.get("t3_earned", 0),
        ])

    for key, ibc in ibc_campuses.items():
        if key not in seen:
            w.writerow([
                key, key, "IBC only",
                0, 0, 0,
                ibc.get("attempts", 0), ibc.get("earned", 0),
                ibc.get("t1_earned", 0), ibc.get("t2_earned", 0), ibc.get("t3_earned", 0),
            ])

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=cte_dashboard_export.csv"},
    )


if FRONTEND_DIST.is_dir():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        file_path = FRONTEND_DIST / full_path
        if file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(FRONTEND_DIST / "index.html"))
