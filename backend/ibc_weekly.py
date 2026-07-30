"""Parse IBC Tracker Weekly Report workbook into dashboard-ready summaries.

Primary sheets used:
  - CCMR Report — campus → certification → teacher hierarchy (earned / failed)
  - IBC Totals — district cert earned totals by grade
  - Raw Data — PEIMS-style attempts (result 1=earned) for code→tier enrichment

Tiers come from Public/CTE Weekly Report/TeaAfSeed.js (via IBC Tiers helpers),
with name/code matching. Unmatched certs are labeled \"No tier\".
"""

from __future__ import annotations

import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from campus_norm import campus_key

ENROLL_ROOT = Path(__file__).resolve().parents[2]
IBC_ROOT = Path(__file__).resolve().parents[3] / "IBC Tiers"
PUBLIC = Path(__file__).resolve().parents[3]

TRACKER_CANDIDATES = [
    ENROLL_ROOT / "Resources" / "IBC Tracker Weekly Report 2025-2026.xlsx",
    IBC_ROOT / "Resources" / "IBC Tracker Weekly Report 2025-2026.xlsx",
    Path.home() / "Downloads" / "IBC Tracker Weekly Report 2025-2026.xlsx",
]

_UPLOADED_TRACKER = os.environ.get("IBC_TRACKER_PATH")
if _UPLOADED_TRACKER:
    TRACKER_CANDIDATES.insert(0, Path(_UPLOADED_TRACKER))

# PEIMS E1733-style in this export: 1=earned, 2=other, 3=failed
RESULT_EARNED = {1, 1.0, "1", "01"}

_cache: Dict[str, Any] = {"data": None, "built_at": None, "path": None, "error": None}


def find_tracker() -> Optional[Path]:
    for p in TRACKER_CANDIDATES:
        if p.is_file():
            return p
    return None


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _tier_bucket(tier: str) -> str:
    t = (tier or "").strip()
    if t == "Tier 1":
        return "t1"
    if t == "Tier 2":
        return "t2"
    if t == "Tier 3":
        return "t3"
    return "none"


def _load_tier_lookups() -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    """Return (code→tier, name_norm→tier, code→name)."""
    if str(IBC_ROOT) not in sys.path:
        sys.path.insert(0, str(IBC_ROOT))
    try:
        from build_tier1_pos_mapping import TEA_SEED_PATH, load_tea_seed  # noqa: WPS433

        _t1, tier_by_id, names_by_id = load_tea_seed(TEA_SEED_PATH)
        by_name: Dict[str, str] = {}
        for code, name in names_by_id.items():
            by_name[_norm(name)] = tier_by_id.get(code, "Unknown")
            short = re.sub(
                r"^(ase\s+entry[-\s]?level\s+automobile\s+|ase\s+)",
                "ase",
                name,
                flags=re.I,
            )
            by_name.setdefault(_norm(short), tier_by_id.get(code, "Unknown"))
        return tier_by_id, by_name, names_by_id
    except Exception:
        return {}, {}, {}


def _resolve_tier(
    name: str,
    code: str,
    code_tiers: Dict[str, str],
    name_tiers: Dict[str, str],
) -> str:
    if code:
        c = str(code).strip()
        if c.isdigit():
            c = str(int(float(c)))
        if c in code_tiers:
            return code_tiers[c]
    nn = _norm(name)
    if nn in name_tiers:
        return name_tiers[nn]
    # containment match (tracker names often truncated)
    for k, tier in name_tiers.items():
        if not k or len(k) < 8:
            continue
        if nn in k or k in nn:
            return tier
    return "No tier"


def _parse_ccmr(ws, code_tiers, name_tiers) -> Dict[str, dict]:
    campuses: Dict[str, dict] = {}
    current_campus: Optional[str] = None
    current_cert: Optional[dict] = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        raw = row[0]
        if raw is None:
            continue
        text = str(raw)
        if not text.strip():
            continue
        lead = len(text) - len(text.lstrip(" "))
        label = text.strip()
        earned = float(row[1] or 0)
        failed = float(row[3] or 0)
        success = row[4]
        projected = float(row[6] or 0) if len(row) > 6 and row[6] is not None else 0.0

        if lead == 0:
            key = campus_key(label)
            campuses[key] = {
                "campus_key": key,
                "display_name": label,
                "earned": int(earned),
                "failed": int(failed),
                "projected": int(projected),
                "success_rate": round(float(success) * 100, 1)
                if isinstance(success, (int, float))
                else 0.0,
                "certs": [],
                "by_tier": Counter(),
            }
            current_campus = key
            current_cert = None
        elif lead == 2 and current_campus:
            tier = _resolve_tier(label, "", code_tiers, name_tiers)
            bucket = _tier_bucket(tier)
            cert = {
                "name": label,
                "earned": int(earned),
                "failed": int(failed),
                "attempts": int(earned + failed),
                "pass_rate": round(100.0 * earned / (earned + failed), 1)
                if (earned + failed)
                else 0.0,
                "projected": int(projected),
                "tier": tier,
                "tier_key": bucket,
                "teachers": [],
            }
            campuses[current_campus]["certs"].append(cert)
            campuses[current_campus]["by_tier"][bucket] += int(earned)
            current_cert = cert
        elif lead >= 4 and current_cert is not None:
            current_cert["teachers"].append(
                {
                    "name": label,
                    "earned": int(earned),
                    "failed": int(failed),
                    "pass_rate": round(100.0 * earned / (earned + failed), 1)
                    if (earned + failed)
                    else 0.0,
                }
            )

    # Sort certs: earned desc, then name
    for d in campuses.values():
        d["certs"].sort(key=lambda c: (-c["earned"], c["name"]))
        d["by_tier"] = dict(d["by_tier"])
        d["t1_earned"] = int(d["by_tier"].get("t1", 0))
        d["t2_earned"] = int(d["by_tier"].get("t2", 0))
        d["t3_earned"] = int(d["by_tier"].get("t3", 0))
        d["none_earned"] = int(d["by_tier"].get("none", 0))
    return campuses


def _parse_ibc_totals(ws, code_tiers, name_tiers) -> List[dict]:
    out: List[dict] = []
    title = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            title = str(row[0] or "") if row else ""
            continue
        if i == 1:
            continue
        name = row[0]
        if not name:
            continue
        name_s = str(name).strip()
        if not name_s or name_s.upper() in {"TOTAL", "TOTALS", "GRAND TOTAL"}:
            continue
        earned = int(float(row[1] or 0))
        projected = int(float(row[2] or 0))
        tier = _resolve_tier(name_s, "", code_tiers, name_tiers)
        out.append(
            {
                "name": name_s,
                "earned": earned,
                "projected": projected,
                "g12": int(float(row[3] or 0)),
                "g11": int(float(row[4] or 0)),
                "g10": int(float(row[5] or 0)),
                "g9": int(float(row[6] or 0)),
                "tier": tier,
                "tier_key": _tier_bucket(tier),
            }
        )
    out.sort(key=lambda c: (-c["earned"], c["name"]))
    return out, title


def _enrich_codes_from_raw(ws, code_tiers, name_tiers) -> Dict[str, str]:
    """Map normalized cert name → TEA code from Raw Data for better tier hits."""
    name_to_code: Dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code = row[8]
        name = row[11]
        if not name:
            continue
        if code is None or code == "":
            continue
        c = str(int(float(code))) if isinstance(code, (int, float)) or str(code).replace(".", "", 1).isdigit() else str(code).strip()
        name_to_code[_norm(str(name))] = c
    # push into name_tiers via codes
    for nn, code in name_to_code.items():
        if code in code_tiers and nn not in name_tiers:
            name_tiers[nn] = code_tiers[code]
    return name_to_code


def load_weekly(force: bool = False) -> Dict[str, Any]:
    global _cache
    if _cache["data"] is not None and not force:
        return _cache["data"]

    path = find_tracker()
    if not path:
        raise FileNotFoundError(
            "IBC Tracker Weekly Report not found. Place "
            "'IBC Tracker Weekly Report 2025-2026.xlsx' in Enrollment Data/Resources/"
        )

    try:
        code_tiers, name_tiers, _names = _load_tier_lookups()
        wb = load_workbook(path, read_only=True, data_only=True)

        if "Raw Data" in wb.sheetnames:
            _enrich_codes_from_raw(wb["Raw Data"], code_tiers, name_tiers)

        campuses = _parse_ccmr(wb["CCMR Report"], code_tiers, name_tiers)
        totals, title = _parse_ibc_totals(wb["IBC Totals"], code_tiers, name_tiers)
        wb.close()

        tier_counts = Counter()
        for c in totals:
            tier_counts[c["tier_key"]] += c["earned"]

        data = {
            "source": str(path),
            "title": title or "IBC Tracker Weekly Report",
            "built_at": datetime.now().isoformat(timespec="seconds"),
            "campuses": sorted(campuses.values(), key=lambda c: c["display_name"]),
            "campus_by_key": campuses,
            "cert_totals": totals,
            "district": {
                "earned": sum(c["earned"] for c in totals),
                "projected": sum(c["projected"] for c in totals),
                "t1": int(tier_counts.get("t1", 0)),
                "t2": int(tier_counts.get("t2", 0)),
                "t3": int(tier_counts.get("t3", 0)),
                "none": int(tier_counts.get("none", 0)),
                "cert_count": len(totals),
                "campus_count": len(campuses),
            },
        }
        _cache = {
            "data": data,
            "built_at": data["built_at"],
            "path": str(path),
            "error": None,
        }
        return data
    except Exception as exc:  # noqa: BLE001
        _cache["error"] = str(exc)
        _cache["data"] = None
        raise


def weekly_status() -> Dict[str, Any]:
    path = find_tracker()
    return {
        "tracker_found": path is not None,
        "tracker_path": str(path) if path else None,
        "loaded": _cache["data"] is not None,
        "built_at": _cache.get("built_at"),
        "error": _cache.get("error"),
    }


def campus_weekly(key: str) -> Optional[dict]:
    data = load_weekly()
    k = campus_key(key)
    hit = data["campus_by_key"].get(k)
    if hit:
        return hit
    # fuzzy display match
    for ck, row in data["campus_by_key"].items():
        if ck == k or row["display_name"].lower() == key.lower():
            return row
        if campus_key(row["display_name"]) == k:
            return row
    return None
